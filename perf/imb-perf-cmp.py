#!/usr/bin/env python3

"""
**********************************************************************
  Copyright(c) 2024, Intel Corporation All rights reserved.

  Redistribution and use in source and binary forms, with or without
  modification, are permitted provided that the following conditions
  are met:
    * Redistributions of source code must retain the above copyright
      notice, this list of conditions and the following disclaimer.
    * Redistributions in binary form must reproduce the above copyright
      notice, this list of conditions and the following disclaimer in
      the documentation and/or other materials provided with the
      distribution.
    * Neither the name of Intel Corporation nor the names of its
      contributors may be used to endorse or promote products derived
      from this software without specific prior written permission.

  THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS
  "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT
  LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR
  A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT
  OWNER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL,
  SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT
  LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE,
  DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY
  THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
  (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE
  OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
**********************************************************************
"""

import argparse
import sys

try:
    import pandas as pd
except ImportError:
    sys.exit("Pandas module not available")

# Try to import openpyxl for Excel spreadsheet creation
try:
    from openpyxl.styles import PatternFill

    openpyxl_available = True
except ImportError:
    openpyxl_available = False


def read_file(file_path):
    # Read the file into a pandas DataFrame and handle errors
    try:
        df = pd.read_csv(file_path, sep="\\s+")
    except pd.errors.ParserError:
        sys.exit(f"Error: Unable to parse file {file_path}")

    return df


def calculate_slope_intercept_diff(df):
    # Extract the SLOPE and INTERCEPT columns
    base_slope = df["SLOPE_BASE"]
    base_intercept = df["INTERCEPT_BASE"]
    compare_slope = df["SLOPE_COMPARE"]
    compare_intercept = df["INTERCEPT_COMPARE"]

    # Calculate the percentage differences
    df["SLOPE_DIFF_%"] = ((compare_slope - base_slope) / base_slope) * 100
    df["INTERCEPT_DIFF_%"] = ((compare_intercept - base_intercept) / base_intercept) * 100

    # Round the values to 2 decimal places
    df = df.round({"SLOPE_DIFF_%": 2, "INTERCEPT_DIFF_%": 2})

    return df


def generate_excel(result_df, threshold, excel_file_name):
    if not openpyxl_available:
        print("Warning: openpyxl module is not available. Excel spreadsheet will not be generated.")
        return

    # Create a Excel writer
    writer = pd.ExcelWriter(excel_file_name, engine="openpyxl")

    # Convert the dataframe to an openpyxl Excel object
    result_df.to_excel(writer, sheet_name="results", index=False)

    # Get the openpyxl workbook and worksheet objects
    worksheet = writer.sheets["results"]

    # Define a format for cells that exceed the threshold
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Iterate through the DataFrame and apply the format
    for idx, row in result_df.iterrows():
        if row["SLOPE_DIFF_%"] > threshold or row["INTERCEPT_DIFF_%"] > threshold:
            for col_num in range(len(row)):
                cell = worksheet.cell(row=idx + 2, column=col_num + 1)
                cell.fill = red_fill
        elif row["SLOPE_DIFF_%"] < -threshold or row["INTERCEPT_DIFF_%"] < -threshold:
            for col_num in range(len(row)):
                cell = worksheet.cell(row=idx + 2, column=col_num + 1)
                cell.fill = green_fill

    # Save the Excel file
    writer.close()


def filter_common_rows(df1, df2):
    merged_df = pd.merge(
        df1, df2, how="outer", on=["ARCH", "CIPHER", "HASH", "DIR", "KEYSZ"], indicator=True
    )
    common_rows_df = pd.merge(df1, df2, how="inner", on=["ARCH", "CIPHER", "HASH", "DIR", "KEYSZ"])
    uncommon_rows_df = merged_df[merged_df["_merge"] != "both"]

    # Remove unnecessary columns
    common_rows_df = common_rows_df.drop(columns=["NO_x", "NO_y"])
    uncommon_rows_df = uncommon_rows_df.drop(columns=["NO_x", "NO_y", "_merge"])

    # Rename columns
    common_rows_df = common_rows_df.rename(
        columns={
            "SLOPE_A_x": "SLOPE_BASE",
            "INTERCEPT_A_x": "INTERCEPT_BASE",
            "SLOPE_A_y": "SLOPE_COMPARE",
            "INTERCEPT_A_y": "INTERCEPT_COMPARE",
        }
    )
    uncommon_rows_df = uncommon_rows_df.rename(
        columns={
            "SLOPE_A_x": "SLOPE_BASE",
            "INTERCEPT_A_x": "INTERCEPT_BASE",
            "SLOPE_A_y": "SLOPE_COMPARE",
            "INTERCEPT_A_y": "INTERCEPT_COMPARE",
        }
    )

    return common_rows_df, uncommon_rows_df


def main():
    parser = argparse.ArgumentParser(
        description="Compare performance diff tool output files. "
        "The tool compares the slope and intercept values of the compare file "
        "against the baseline file and checks if the differences exceed a percentage threshold. "
        "Print any results where ((compare - baseline) / baseline) * 100 > threshold %."
    )
    parser.add_argument(
        "baseline_file", type=str, help="File containing baseline performance numbers"
    )
    parser.add_argument("compare_file", type=str, help="File to compare against baseline_file")
    parser.add_argument(
        "-t",
        "--threshold",
        type=int,
        default=5,
        help="Threshold percentage to check slope and intercept against (default: 5)",
    )
    parser.add_argument(
        "-e",
        "--generate_excel",
        action="store_true",
        default=False,
        help="Generate an Excel spreadsheet with results",
    )
    parser.add_argument(
        "-f",
        "--excel_file_name",
        type=str,
        default="ipsec-perf-cmp_result.xlsx",
        help="Name of the Excel file to generate (default: ipsec-perf-cmp_result.xlsx)",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        default=False,
        help="Print verbose output",
    )

    args = parser.parse_args()

    baseline_file = args.baseline_file
    compare_file = args.compare_file
    threshold = args.threshold
    generate_excel_flag = args.generate_excel
    excel_file_name = args.excel_file_name
    verbose_flag = args.verbose

    # Read the files
    baseline_df = read_file(baseline_file)
    compare_df = read_file(compare_file)

    # Set pandas options to display all rows and columns
    pd.set_option("display.max_rows", None)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", None)
    pd.set_option("display.max_colwidth", None)

    # Filter the common and uncommon rows
    common_df, uncommon_df = filter_common_rows(baseline_df, compare_df)

    # Print the rows that are not common in both files
    if not uncommon_df.empty and verbose_flag:
        print("The following rows are not common in both files:")
        print(uncommon_df)
        print()

    # Compare slope and intercept values
    result_df = calculate_slope_intercept_diff(common_df)

    # Filter the rows where SLOPE_DIFF_% or INTERCEPT_DIFF_% exceed the threshold
    exceeding_threshold_df = result_df[
        (result_df["SLOPE_DIFF_%"] > threshold) | (result_df["INTERCEPT_DIFF_%"] > threshold)
    ]

    # Generate Excel file if the flag is set
    if generate_excel_flag:
        generate_excel(result_df, threshold, excel_file_name)

    # Exit with an error if exceeding_threshold_df is not empty
    if not exceeding_threshold_df.empty:
        print("Differences found exceeding the threshold:")
        print(exceeding_threshold_df)
        sys.exit(1)
    else:
        print("No differences found.")


if __name__ == "__main__":
    main()
